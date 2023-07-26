import  openpyxl

ruta_archivo = openpyxl.load_workbook("C:/Users/endor/OneDrive/Documentos/introduccion_robot/files/Data.xlsx")

def Numero_Filas(hoja):
    accion = ruta_archivo[hoja]
    return  accion.max_column

def Dato_Columna(hoja, fila, columa):
    accion = ruta_archivo[hoja]
    columa = accion.cell(fila,columa)
    return  columa.value


print("El numero de la fila -> "+ str(Numero_Filas("Hoja1")))
print("El valor de la columna -> "+ str(Dato_Columna("Hoja1",1,2)))

